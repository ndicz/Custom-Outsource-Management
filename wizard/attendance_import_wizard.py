from odoo import fields, models
from odoo.exceptions import UserError
import base64
import csv
from io import StringIO, BytesIO


class AttendanceImportWizard(models.TransientModel):
    _name = 'attendance.import.wizard'
    _description = 'Import Absensi CSV/XLSX'

    payroll_period_id = fields.Many2one('hr.payroll.period', string='Periode Payroll', required=True)
    file_data = fields.Binary(string='File Absensi', required=True)
    file_name = fields.Char(string='Nama File', required=True)
    file_type = fields.Selection([
        ('csv', 'CSV'),
        ('xlsx', 'Excel (.xlsx)'),
    ], string='Tipe File', default='csv', required=True)
    csv_delimiter = fields.Selection([
        (',', 'Koma (,)'),
        (';', 'Titik Koma (;)'),
    ], string='Pemisah CSV', default=',')

    default_absence_deduction_rate = fields.Float(
        string='Default Potongan Alfa/Hari',
        default=0.0,
        help='Dipakai jika kolom absence_deduction_rate tidak ada di file.'
    )
    default_holiday_work_rate = fields.Float(
        string='Default Insentif Libur/Hari',
        default=0.0,
        help='Dipakai jika kolom holiday_work_rate tidak ada di file.'
    )

    result_message = fields.Text(string='Hasil Import', readonly=True)

    def action_import_attendance(self):
        self.ensure_one()

        if not self.file_data:
            raise UserError('File absensi wajib diisi.')

        rows = self._load_rows()
        if not rows:
            raise UserError('File tidak berisi data absensi.')

        slips = self.payroll_period_id.payslip_ids
        if not slips:
            raise UserError('Belum ada payslip di periode ini. Jalankan Auto-Generate Slips dulu.')

        updated = 0
        not_found = []

        for idx, row in enumerate(rows, start=2):
            slip = self._find_payslip(slips, row)
            if not slip:
                key = row.get('employee_name') or row.get('employee') or row.get('barcode') or str(idx)
                not_found.append(key)
                continue

            days_worked = int(float(row.get('days_worked', 0) or 0))
            absent_days = int(float(row.get('absent_days', 0) or 0))
            public_holiday_days = int(float(row.get('public_holiday_days', 0) or 0))
            holiday_work_days = float(row.get('holiday_work_days', 0) or 0)
            overtime_hours = float(row.get('overtime_hours', 0) or 0)
            absence_rate = float(
                row.get('absence_deduction_rate', self.default_absence_deduction_rate or 0.0) or 0.0
            )
            holiday_rate = float(
                row.get('holiday_work_rate', self.default_holiday_work_rate or 0.0) or 0.0
            )

            if absence_rate <= 0 and slip.basic_salary:
                absence_rate = slip.basic_salary / 20
            if holiday_rate <= 0 and slip.basic_salary:
                holiday_rate = slip.basic_salary / 20

            slip.write({
                'days_worked': max(days_worked, 0),
                'absent_days': max(absent_days, 0),
                'public_holiday_days': max(public_holiday_days, 0),
                'holiday_work_days': max(holiday_work_days, 0.0),
                'overtime_hours': max(overtime_hours, 0.0),
                'absence_deduction_rate': max(absence_rate, 0.0),
                'holiday_work_rate': max(holiday_rate, 0.0),
                'attendance_source': 'import',
            })
            updated += 1

        message = f'Import selesai. Berhasil update {updated} slip.'
        if not_found:
            preview = ', '.join(not_found[:10])
            message += f' Data karyawan tidak ditemukan ({len(not_found)}): {preview}'

        self.write({'result_message': message})

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'attendance.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def _find_payslip(self, slips, row):
        employee_id = row.get('employee_id')
        if employee_id:
            try:
                return slips.filtered(lambda s: s.employee_id.id == int(float(employee_id)))[:1]
            except Exception:
                pass

        barcode = (row.get('barcode') or '').strip()
        if barcode:
            found = slips.filtered(lambda s: (s.employee_id.barcode or '').strip().lower() == barcode.lower())
            if found:
                return found[:1]

        email = (row.get('work_email') or '').strip()
        if email:
            found = slips.filtered(lambda s: (s.employee_id.work_email or '').strip().lower() == email.lower())
            if found:
                return found[:1]

        employee_name = (row.get('employee_name') or row.get('employee') or '').strip()
        if employee_name:
            found = slips.filtered(lambda s: (s.employee_id.name or '').strip().lower() == employee_name.lower())
            if found:
                return found[:1]

        return False

    def _load_rows(self):
        raw = base64.b64decode(self.file_data)
        if self.file_type == 'xlsx':
            return self._load_xlsx_rows(raw)
        return self._load_csv_rows(raw)

    def _load_csv_rows(self, raw):
        text = raw.decode('utf-8-sig')
        reader = csv.DictReader(StringIO(text), delimiter=self.csv_delimiter or ',')
        return [self._normalize_row(row) for row in reader if row]

    def _load_xlsx_rows(self, raw):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise UserError('Import Excel butuh package openpyxl di environment Odoo.') from exc

        workbook = load_workbook(filename=BytesIO(raw), data_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        data = []
        for values in rows[1:]:
            if not any(values):
                continue
            row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
            data.append(self._normalize_row(row))
        return data

    def _normalize_row(self, row):
        normalized = {}
        for key, value in row.items():
            key_norm = (key or '').strip().lower()
            normalized[key_norm] = value
        return normalized
